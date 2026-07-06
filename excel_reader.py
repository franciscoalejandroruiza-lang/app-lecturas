"""
excel_reader.py
Lee la hoja "OK" del Excel de consumo mensual, que crece agregando columnas
nuevas cada mes (Carta B&N, Oficio B&N, Carta Color, Oficio Color, y
usualmente Digitalización). Los encabezados están escritos a mano y tienen
inconsistencias reales (bloques de 4 columnas sin digitalización, meses
duplicados o mal copiados, columnas sueltas de "FECHA Y HORA"), así que los
bloques se detectan por el CONTENIDO de cada encabezado (qué campo es) y no
por posición fija, agrupando cada bloque desde una columna "Carta B&N" hasta
la siguiente. Aun así, SIEMPRE se le pide confirmación al usuario en la UI.
"""
import re
import openpyxl

EXCEL_MES_CAMPOS = ["carta_bn", "oficio_bn", "carta_color", "oficio_color", "digitalizacion"]

MESES = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO",
         "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
MESES_ABBR = {"ENE": "ENERO", "FEB": "FEBRERO", "MAR": "MARZO", "ABR": "ABRIL",
              "MAY": "MAYO", "JUN": "JUNIO", "JUL": "JULIO", "AGO": "AGOSTO",
              "SEP": "SEPTIEMBRE", "OCT": "OCTUBRE", "NOV": "NOVIEMBRE", "DIC": "DICIEMBRE"}


def load_ok_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'OK' not in wb.sheetnames:
        raise ValueError("El archivo no tiene una hoja llamada 'OK'.")
    return wb['OK']


def _classify_col(text):
    """Clasifica una columna de encabezado según su contenido de texto."""
    if not text:
        return None
    t = str(text).upper().replace('\n', ' ')
    if 'FECHA' in t and 'HORA' in t:
        return 'fecha'
    if 'DIGITALIZ' in t or 'DIFITALIZ' in t:  # typo real presente en el archivo
        return 'digitalizacion'
    has_carta = 'CARTA' in t
    has_oficio = 'OFICIO' in t
    has_color = 'COLOR' in t
    if has_carta and has_color:
        return 'carta_color'
    if has_oficio and has_color:
        return 'oficio_color'
    if has_carta:
        return 'carta_bn'
    if has_oficio:
        return 'oficio_bn'
    return None


def _extract_month(text):
    if not text:
        return None
    t = str(text).upper().replace('\n', ' ')
    for full in MESES:
        if full in t:
            return full
    for abbr, full in MESES_ABBR.items():
        if re.search(rf'\b{abbr}\b', t):
            return full
    return None


def detect_month_blocks(ws):
    """Detecta bloques de columnas de lectura mensual agrupando por tipo de
    campo (no por posición fija de 5), desde la col 7 en adelante. Cada
    bloque puede traer además una columna de 'Fecha y Hora' asociada."""
    max_col = ws.max_column
    real_max = 6
    for c in range(7, min(max_col, 2000) + 1):
        if ws.cell(row=2, column=c).value not in (None, ''):
            real_max = c

    col_types = {}
    for c in range(7, real_max + 1):
        col_types[c] = _classify_col(ws.cell(row=2, column=c).value)

    blocks_raw = []
    current_cols = []
    current_fecha = None
    for c in range(7, real_max + 1):
        t = col_types[c]
        if t is None:
            continue
        if t == 'carta_bn' and current_cols:
            blocks_raw.append((current_cols, current_fecha))
            current_cols = []
            current_fecha = None
        if t == 'fecha':
            current_fecha = c
        else:
            current_cols.append(c)
    if current_cols:
        blocks_raw.append((current_cols, current_fecha))

    from openpyxl.utils import get_column_letter
    result = []
    for i, (cols, fecha_col) in enumerate(blocks_raw):
        month = None
        for c in cols:
            month = month or _extract_month(ws.cell(row=1, column=c).value)
            month = month or _extract_month(ws.cell(row=2, column=c).value)
        col_range = f"{get_column_letter(cols[0])}:{get_column_letter(cols[-1])}"
        label = f"{month} ({col_range})" if month else f"Bloque {i+1} sin nombre de mes ({col_range})"
        result.append({'cols': cols, 'fecha_col': fecha_col, 'label': label, 'col_range': col_range, 'month': month})
    return result


def read_block_values(ws, block):
    """Regresa { serie: {carta_bn, oficio_bn, carta_color, oficio_color, digitalizacion, fecha_hora} }
    Si el bloque tiene menos de 5 columnas (p.ej. no se capturó digitalización
    ese mes), los campos faltantes quedan en None."""
    cols = block['cols']
    fecha_col = block.get('fecha_col')
    out = {}
    for r in range(3, ws.max_row + 1):
        serie = ws.cell(row=r, column=6).value
        if not serie:
            continue
        serie = str(serie).strip()
        vals = {}
        for idx, campo in enumerate(EXCEL_MES_CAMPOS):
            if idx < len(cols):
                v = ws.cell(row=r, column=cols[idx]).value
                vals[campo] = v
            else:
                vals[campo] = None
        if fecha_col:
            v = ws.cell(row=r, column=fecha_col).value
            vals['fecha_hora'] = str(v).strip() if v else None
        else:
            vals['fecha_hora'] = None
        out[serie] = vals
    return out


def read_unidad_map(ws):
    """serie -> nombre de unidad (columna B), para agrupar por ubicación."""
    out = {}
    for r in range(3, ws.max_row + 1):
        serie = ws.cell(row=r, column=6).value
        unidad = ws.cell(row=r, column=2).value
        if serie:
            out[str(serie).strip()] = (unidad or '').strip()
    return out
