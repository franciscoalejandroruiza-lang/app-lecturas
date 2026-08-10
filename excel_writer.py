"""
excel_writer.py
Escribe los valores capturados (ya calculados) de vuelta en la hoja "OK" del
Excel de consumo, ya sea actualizando un bloque de mes existente o creando
uno nuevo al final. No usa fórmulas de Excel, solo valores (igual que el
archivo original).
"""
from openpyxl.utils import get_column_letter

from excel_reader import EXCEL_MES_CAMPOS


def find_row_for_serie(ws, serie):
    serie = str(serie).strip()
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=6).value
        if v and str(v).strip() == serie:
            return r
    return None


def write_values_to_block(ws, serie, block, values, fecha_hora=None):
    """Escribe carta_bn/oficio_bn/carta_color/oficio_color/digitalizacion (lo
    que venga en `values`, None se deja intacto) en las columnas del bloque,
    en la fila correspondiente a `serie`. Regresa (ok, mensaje)."""
    row = find_row_for_serie(ws, serie)
    if row is None:
        return False, f"No se encontró la serie {serie} en la hoja 'OK' (columna F)."

    cols = block["cols"]
    for idx, campo in enumerate(EXCEL_MES_CAMPOS):
        if idx >= len(cols):
            continue
        val = values.get(campo)
        if val is not None:
            ws.cell(row=row, column=cols[idx], value=val)

    if fecha_hora and block.get("fecha_col"):
        ws.cell(row=row, column=block["fecha_col"], value=fecha_hora)

    return True, "OK"


def _find_real_last_col(ws):
    """ws.max_column puede regresar columnas 'fantasma' con formato pero sin
    datos (hasta el límite de Excel). Se busca la última columna con
    encabezado real en la fila 2, igual que en detect_month_blocks."""
    real_max = 6
    for c in range(7, min(ws.max_column, 2000) + 1):
        if ws.cell(row=2, column=c).value not in (None, ''):
            real_max = c
    return real_max


def create_new_month_block(ws, mes_nombre):
    """Agrega un bloque nuevo de 6 columnas (Carta B&N, Oficio B&N, Carta
    Color, Oficio Color, Digitalización, Fecha y Hora) al final de la hoja,
    con encabezados en la fila 2. Regresa el dict de bloque (mismo formato
    que detect_month_blocks) para usarlo de inmediato."""
    start_col = _find_real_last_col(ws) + 1
    headers = [
        f"LECTURA INICIAL CARTA B&N {mes_nombre}",
        f"LECTURA INICIAL OFICIO B&N {mes_nombre}",
        f"LECTURA INICIAL CARTA COLOR {mes_nombre}",
        f"LECTURA INICIAL OFICIO COLOR {mes_nombre}",
        f"DIGITALIZACION {mes_nombre}",
        "FECHA Y HORA",
    ]
    for i, h in enumerate(headers):
        ws.cell(row=2, column=start_col + i, value=h)
    ws.cell(row=1, column=start_col, value=mes_nombre)

    cols = list(range(start_col, start_col + 5))
    fecha_col = start_col + 5
    col_range = f"{get_column_letter(cols[0])}:{get_column_letter(fecha_col)}"
    return {
        "cols": cols, "fecha_col": fecha_col,
        "label": f"{mes_nombre} ({col_range})", "col_range": col_range, "month": mes_nombre,
    }
